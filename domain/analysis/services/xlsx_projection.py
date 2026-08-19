from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone as dt_timezone
from typing import Any, Optional

from openpyxl.utils import get_column_letter

from domain.analysis.models import (
    DigitalAnalysisResult,
    MultiTagAnalysisResult,
    SegmentKind,
    TagAnalysisResult,
)
from domain.analysis.services._digital import format_duration


@dataclass(frozen=True)
class XlsxCellStyle:
    bg_color: Optional[str] = None
    font_color: Optional[str] = None
    bold: bool = False
    align: Optional[str] = None
    border: bool = False
    number_format: Optional[str] = None
    wrap_text: bool = False


@dataclass(frozen=True)
class XlsxSheet:
    name: str
    columns: list[str]
    rows: list[list[Any]]
    warnings: list[str] = field(default_factory=list)
    is_presentation: bool = False
    freeze_panes: Optional[str] = None
    column_widths: dict[int, float] = field(default_factory=dict)
    merges: list[str] = field(default_factory=list)
    cell_styles: dict[tuple[int, int], XlsxCellStyle] = field(default_factory=dict)
    is_active: bool = False


_STATUS_DESCRIPTIONS = {
    "complete": "Janela com cobertura válida e estados digitais reconhecidos.",
    "no_transitions": "A tag permaneceu no mesmo estado conhecido durante toda ou quase toda a janela.",
    "partial_coverage": "Parte da janela não pôde ser reconstruída.",
    "invalid_digital_values": "Não houve segmento KNOWN suficiente na janela.",
    "no_data": "Não foram encontrados dados suficientes para reconstruir a janela.",
}

_DIAGNOSIS_TEMPLATES = {
    "complete": "Janela com cobertura válida e estados digitais reconhecidos.",
    "no_transitions": "A tag permaneceu no mesmo estado conhecido durante toda ou quase toda a janela.",
    "partial_coverage": "Parte da janela não pôde ser reconstruída. Veja aba Linha_do_Tempo.",
    "invalid_digital_values": "Não houve segmento KNOWN suficiente na janela.",
    "no_data": "Não foram encontrados dados suficientes para reconstruir a janela.",
}

_WARNING_RECOMMENDATIONS = {
    "UNKNOWN_DIGITAL_VALUES": "Compare a aba Valores_Desconhecidos com a configuração atual do Digital Set. Valores fora do Digital Set podem indicar alteração histórica do set, escrita inesperada ou divergência de configuração.",
    "NO_KNOWN_DIGITAL_STATES": "Investigue Valores_Desconhecidos e Linha_do_Tempo.",
    "PARTIAL_TIMELINE_COVERAGE": "Verifique a disponibilidade de dados na aba Linha_do_Tempo.",
    "SEED_BAD_QUALITY": "O período inicial pode estar classificado como BAD por causa do seed.",
    "BAD_QUALITY_COVERAGE": "Investigue a integridade da coleta no PI nesta faixa.",
}

_STATE_COLORS = [
    "4CAF50", "2196F3", "FF9800", "9C27B0", "00BCD4",
    "E91E63", "795548", "607D8B", "3F51B5", "009688",
    "8BC34A", "FFC107"
]

_QUALITY_COLORS = {
    "KNOWN": "2E7D32",
    "GOOD": "2E7D32",
    "BAD": "D32F2F",
    "UNKNOWN": "ED6C02",
    "NULL": "9E9E9E",
    "UNCOVERED": "D1C4E9",
    "MIXED": "757575",
}

_MIXED_COLOR = "757575"
DEFAULT_PRESENTATION_TZ = "America/Sao_Paulo"


def _parse_datetime(val: str | datetime) -> datetime:
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(val)
    except Exception:
        from datetime import timezone as dt_tz
        return datetime.now(dt_tz.utc)


def _format_presentation_time(val: str | datetime | None, tz_name: str = DEFAULT_PRESENTATION_TZ) -> str:
    if val is None:
        return ""
    dt = _parse_datetime(val)
    if dt.tzinfo is None:
        from datetime import timezone as dt_tz
        dt = dt.replace(tzinfo=dt_tz.utc)
    try:
        from zoneinfo import ZoneInfo
        local_dt = dt.astimezone(ZoneInfo(tz_name))
        return local_dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return dt.strftime("%Y-%m-%d %H:%M:%S")


def calculate_bucket_count(window_seconds: float, segment_count: int) -> int:
    if window_seconds <= 0:
        return 96
    target = max(96, min(240, segment_count if segment_count > 0 else 96))
    return target


class XlsxAnalysisProjection:
    def project(self, multi: MultiTagAnalysisResult) -> list[XlsxSheet]:
        if self._is_all_numeric(multi):
            return self._project_numeric_only(multi)
        if self._is_all_digital(multi):
            return self._project_digital_only(multi)
        return self._project_mixed(multi)

    def _is_all_numeric(self, multi: MultiTagAnalysisResult) -> bool:
        return all(r.metadata.point_type == "numeric" for r in multi.results)

    def _is_all_digital(self, multi: MultiTagAnalysisResult) -> bool:
        return all(r.metadata.point_type == "digital" for r in multi.results)

    # ------------------------------------------------------------------
    # Numeric-only (preservado do baseline)
    # ------------------------------------------------------------------

    def _project_numeric_only(self, multi: MultiTagAnalysisResult) -> list[XlsxSheet]:
        return [
            self._resumo(multi),
            self._qualidade(multi),
            self._estatisticas(multi),
            self._recorded(multi),
            self._interpolated_5m(multi),
            self._gaps(multi),
            self._spikes(multi),
            self._erros_warnings(multi),
            self._metadados(multi),
        ]

    def _resumo(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag",
            "point_type",
            "descriptor",
            "engineering_units",
            "janela_solicitada_inicio",
            "janela_solicitada_fim",
            "janela_efetiva_inicio",
            "janela_efetiva_fim",
            "pontos_retornados",
            "completude",
            "verdict",
        ]
        rows = []
        for r in multi.results:
            verdict = "NÃO APLICÁVEL" if r.digital_analysis is not None else (r.quality.verdict if r.quality else "")
            c = r.completeness
            eff_start = c.effective_start_time if c else multi.period_start
            eff_end = c.effective_end_time if c else multi.period_end
            pts_cnt = c.returned_point_count if c else ""
            comp_str = c.analysis_completeness.value if c else "COMPLETE"

            rows.append([
                r.metadata.tag,
                r.metadata.point_type,
                r.metadata.descriptor,
                r.metadata.engineering_units or "",
                multi.period_start,
                multi.period_end,
                eff_start,
                eff_end,
                pts_cnt,
                comp_str,
                verdict,
            ])
        return XlsxSheet(name="Resumo", columns=columns, rows=rows)

    def _qualidade(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "good_pct", "questionable_pct", "substituted_pct", "zero_pct", "zero_policy", "verdict"]
        rows = []
        for r in multi.results:
            if r.digital_analysis is not None:
                rows.append([
                    r.metadata.tag, None, None, None, None,
                    r.zero_policy_applied, "NÃO APLICÁVEL",
                ])
            elif r.quality is not None:
                rows.append([
                    r.metadata.tag,
                    r.quality.good_pct,
                    r.quality.questionable_pct,
                    r.quality.substituted_pct,
                    r.quality.zero_pct,
                    r.zero_policy_applied,
                    r.quality.verdict,
                ])
            else:
                rows.append([r.metadata.tag, None, None, None, None, r.zero_policy_applied, ""])
        return XlsxSheet(name="Qualidade", columns=columns, rows=rows)

    def _estatisticas(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "count", "min", "max", "mean", "median", "p01", "p99", "stddev_pop", "stddev_sample", "sum", "zero_count"]
        rows = []
        for r in multi.results:
            if r.numeric:
                s = r.numeric
                rows.append([
                    r.metadata.tag, s.count, s.min, s.max, s.mean, s.median,
                    s.p01, s.p99, s.stddev_pop, s.stddev_sample, s.sum, s.zero_count,
                ])
            else:
                rows.append([r.metadata.tag, None, None, None, None, None, None, None, None, None, None, None])
        return XlsxSheet(name="Estatisticas", columns=columns, rows=rows)

    def _recorded(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "timestamp", "value", "good", "questionable", "substituted", "source"]
        rows = []
        return XlsxSheet(name="Recorded", columns=columns, rows=rows)

    def _interpolated_5m(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "timestamp", "value", "good", "questionable", "substituted"]
        rows = []
        has_digital = any(r.metadata.point_type == "digital" for r in multi.results)
        if has_digital and len(multi.results) > 0:
            rows.append(["N/A para tags digitais", "", "", "", "", ""])
        return XlsxSheet(name="Interpolated_5m", columns=columns, rows=rows)

    def _gaps(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "method", "start_ts", "end_ts", "duration_seconds"]
        rows = []
        for r in multi.results:
            for g in r.gaps_interpolated:
                rows.append([r.metadata.tag, "interpolated", g.start_ts, g.end_ts, g.duration_seconds])
            for g in r.gaps_recorded:
                rows.append([r.metadata.tag, "recorded", g.start_ts, g.end_ts, g.duration_seconds])
        return XlsxSheet(name="Gaps", columns=columns, rows=rows)

    def _spikes(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "timestamp", "prev_value", "cur_value", "abs_delta", "rel_delta", "basis"]
        rows = []
        for r in multi.results:
            for s in r.spikes:
                rows.append([
                    r.metadata.tag, s.timestamp, s.previous_value,
                    s.current_value, s.absolute_delta, s.relative_delta,
                    s.detection_basis,
                ])
        return XlsxSheet(name="Spikes", columns=columns, rows=rows)

    def _erros_warnings(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "code", "message", "retryable"]
        rows = []
        for e in multi.errors:
            rows.append([e.tag or "", e.code, e.message, e.retryable])

        for r in multi.results:
            c = r.completeness
            if c is not None and c.truncated:
                rows.append([
                    r.metadata.tag,
                    "POINT_LIMIT_EXCEEDED",
                    (
                        f"Limite de pontos por tag excedido ({c.returned_point_count}/{c.effective_point_limit}). "
                        f"Período analisado: {c.effective_start_time} até {c.effective_end_time}. "
                        f"Período não analisado: {c.unprocessed_start_time} até {c.unprocessed_end_time}."
                    ),
                    False,
                ])

        return XlsxSheet(name="Erros_Warnings", columns=columns, rows=rows)

    def _metadados(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["key", "value"]
        rows = [
            ["tool_name", "generate_pi_tags_analysis_report"],
            ["period_start", multi.period_start],
            ["period_end", multi.period_end],
            ["total_requested", str(multi.total_requested)],
            ["total_processed", str(multi.total_processed)],
            ["overall_completeness", multi.overall_completeness.value if multi.overall_completeness else "COMPLETE"],
        ]

        if multi.results and multi.results[0].completeness:
            c = multi.results[0].completeness
            rows.extend([
                ["configured_point_limit", str(c.configured_point_limit)],
                ["effective_point_limit", str(c.effective_point_limit)],
                ["returned_point_count", str(c.returned_point_count)],
                ["truncation_direction", c.truncation_direction],
                ["overflow_check_performed", str(c.overflow_check_performed)],
            ])
        
        rows.append(["schema_version", "1.0"])
        return XlsxSheet(name="Metadados", columns=columns, rows=rows)

    # ------------------------------------------------------------------
    # Digital-only (Redesenho Orientado à Detecção de Problemas)
    # ------------------------------------------------------------------

    def _project_digital_only(self, multi: MultiTagAnalysisResult) -> list[XlsxSheet]:
        sheets: list[XlsxSheet] = []
        sheets.append(self._visao_geral(multi))

        first_da = next(
            (r.digital_analysis for r in multi.results if r.digital_analysis is not None),
            None,
        )
        if first_da is not None and first_da.timeline_segments:
            sheets.append(self._linha_do_tempo(multi))

        sheets.append(self._resumo_digital(multi))
        sheets.append(self._qualidade_digital(multi))
        sheets.append(self._recorded_digital(multi))

        if first_da is not None:
            if first_da.state_statistics:
                sheets.append(self._estados(multi))
            if first_da.transition_statistics:
                sheets.append(self._transicoes(multi))
            if first_da.unknown_value_statistics:
                sheets.append(self._valores_desconhecidos(multi))
            if first_da.daily_summary:
                sheets.append(self._resumo_diario(multi))
            if first_da.digital_set_snapshot:
                sheets.append(self._digital_set(multi))

        sheets.append(self._erros_warnings(multi))
        sheets.append(self._metadados(multi))
        return sheets

    def _visao_geral(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        from datetime import date, timedelta
        from zoneinfo import ZoneInfo

        rows: list[list[Any]] = []
        cell_styles: dict[tuple[int, int], XlsxCellStyle] = {}
        merges: list[str] = []

        digital_results = [r for r in multi.results if r.metadata.point_type == "digital"]

        w_start_dt = _parse_datetime(multi.period_start)
        w_end_dt = _parse_datetime(multi.period_end)

        try:
            tz = ZoneInfo(DEFAULT_PRESENTATION_TZ)
            w_start_local = w_start_dt.astimezone(tz)
            w_end_local = w_end_dt.astimezone(tz)
        except Exception:
            from datetime import timezone as dt_tz
            tz = dt_tz.utc
            w_start_local = w_start_dt
            w_end_local = w_end_dt

        start_d = w_start_local.date()
        end_d = w_end_local.date()
        if w_end_local.time() == datetime.min.time() and end_d > start_d:
            end_d = end_d - timedelta(days=1)

        days: list[date] = []
        curr_d = start_d
        while curr_d <= end_d:
            days.append(curr_d)
            curr_d += timedelta(days=1)

        num_cols = max(14, len(days) + 1)

        title_style = XlsxCellStyle(bg_color="1A365D", font_color="FFFFFF", bold=True, align="center")
        section_style = XlsxCellStyle(bg_color="2B6CB0", font_color="FFFFFF", bold=True, align="left")
        header_lbl_style = XlsxCellStyle(bg_color="EBF8FF", bold=True, wrap_text=True)
        header_val_style = XlsxCellStyle(wrap_text=True)
        tbl_hdr_style = XlsxCellStyle(bg_color="1A365D", font_color="FFFFFF", bold=True, align="center", border=True, wrap_text=True)
        tbl_val_style = XlsxCellStyle(align="center", border=True, wrap_text=True)
        kpi_card_style = XlsxCellStyle(bg_color="EDF2F7", bold=True, align="center", border=True, wrap_text=True)
        kpi_val_style = XlsxCellStyle(bg_color="FFFFFF", bold=True, align="center", border=True)

        # Row 1: Title Banner
        rows.append(["RELATÓRIO EXECUTIVO - DIAGNÓSTICO E LINHA DO TEMPO DE TAGS DIGITAIS"] + [""] * (num_cols - 1))
        merges.append(f"A1:{get_column_letter(num_cols)}1")
        cell_styles[(1, 1)] = title_style

        # Section 1: Identificação e Status
        rows.append(["1. IDENTIFICAÇÃO E DIAGNÓSTICO EXECUTIVO"] + [""] * (num_cols - 1))
        merges.append(f"A2:{get_column_letter(num_cols)}2")
        cell_styles[(2, 1)] = section_style

        p_start_local_str = _format_presentation_time(multi.period_start)
        p_end_local_str = _format_presentation_time(multi.period_end)

        bad_days: set[date] = set()
        unknown_days: set[date] = set()
        null_days: set[date] = set()
        uncovered_days: set[date] = set()

        first_r = digital_results[0] if digital_results else None
        da_first = first_r.digital_analysis if first_r else None
        qs = da_first.quality_summary if da_first else None
        cov = da_first.coverage if da_first else None

        bad_pts = qs.bad_events if qs else 0
        bad_dur = cov.bad_seconds if cov else 0.0
        first_bad = qs.first_bad_timestamp if qs else None
        last_bad = qs.last_bad_timestamp if qs else None

        for r in digital_results:
            da = r.digital_analysis
            descriptor = r.metadata.descriptor or "N/A"
            digital_set = r.metadata.digital_set or "N/A"
            timeline_segs = da.timeline_segments if da else ()

            state_color_map: dict[str, str] = {}
            if da and da.possible_states:
                for idx, ps in enumerate(da.possible_states):
                    state_color_map[ps.state_name] = _STATE_COLORS[idx % len(_STATE_COLORS)]

            daily_summaries: list[dict[str, Any]] = []
            for d in days:
                d_start = datetime.combine(d, datetime.min.time(), tzinfo=tz)
                d_end = datetime.combine(d + timedelta(days=1), datetime.min.time(), tzinfo=tz)

                slice_start = max(d_start, w_start_local)
                slice_end = min(d_end, w_end_local)
                slice_dur = max(0.0, (slice_end - slice_start).total_seconds())

                is_partial_day = (slice_start > d_start or slice_end < d_end)

                d_bad_dur = 0.0
                d_unknown_dur = 0.0
                d_null_dur = 0.0
                d_uncovered_dur = 0.0
                d_known_dur = 0.0
                d_state_durs: dict[str, float] = {}
                d_kinds: set[str] = set()

                for seg in timeline_segs:
                    s_st = _parse_datetime(seg.start).astimezone(tz)
                    s_en = _parse_datetime(seg.end).astimezone(tz)
                    overlap = max(0.0, (min(s_en, slice_end) - max(s_st, slice_start)).total_seconds())
                    if overlap > 0:
                        k_val = seg.kind.value if hasattr(seg.kind, "value") else str(seg.kind)
                        k_val = k_val.lower()
                        d_kinds.add(k_val)

                        if k_val == "bad":
                            d_bad_dur += overlap
                            bad_days.add(d)
                        elif k_val == "unknown":
                            d_unknown_dur += overlap
                            unknown_days.add(d)
                        elif k_val == "null":
                            d_null_dur += overlap
                            null_days.add(d)
                        elif k_val == "uncovered":
                            d_uncovered_dur += overlap
                            uncovered_days.add(d)
                        elif k_val in ("known", "good"):
                            d_known_dur += overlap
                            st_name = seg.state_name or "KNOWN"
                            d_state_durs[st_name] = d_state_durs.get(st_name, 0.0) + overlap

                if d_bad_dur > 0:
                    prim = "BAD"
                elif d_unknown_dur > 0:
                    prim = "UNKNOWN"
                elif d_null_dur > 0:
                    prim = "NULL"
                elif d_uncovered_dur > 0:
                    prim = "UNCOVERED"
                else:
                    prim = "OK"

                pred_st = "N/A"
                pred_color = "F0F0F0"
                if d_state_durs:
                    max_st = max(d_state_durs.items(), key=lambda x: x[1])
                    pred_st = max_st[0]
                    pred_color = state_color_map.get(pred_st, _STATE_COLORS[0])

                daily_summaries.append({
                    "date": d,
                    "primary": prim,
                    "bad_dur": d_bad_dur,
                    "unknown_dur": d_unknown_dur,
                    "null_dur": d_null_dur,
                    "uncovered_dur": d_uncovered_dur,
                    "known_dur": d_known_dur,
                    "slice_dur": slice_dur,
                    "is_partial": is_partial_day,
                    "pred_state": pred_st,
                    "pred_color": pred_color,
                    "kind_count": len(d_kinds),
                })

            comp_meta = multi.results[0].completeness if multi.results and multi.results[0].completeness else None
            if comp_meta and comp_meta.truncated:
                status_exec = "ANÁLISE PARCIAL (LIMITE DE PONTOS ALCANÇADO)"
            elif bad_days:
                status_exec = "PROBLEMAS DETECTADOS (QUALIDADE BAD)"
            elif unknown_days or null_days or uncovered_days:
                status_exec = "PROBLEMAS DETECTADOS (DESCONHECIDO / FALTA COBERTURA)"
            else:
                status_exec = "SEM PROBLEMAS DETECTADOS"

            if bad_days:
                diag_text = (
                    f"ATENÇÃO: Foram identificados períodos com qualidade BAD em {len(bad_days)} dia(s) da janela analisada, "
                    f"totalizando {format_duration(bad_dur)} ({qs.bad_segment_count if qs else 0} intervalos contínuos e {bad_pts} pontos PI). "
                    f"Primeira ocorrência: {_format_presentation_time(first_bad) or 'N/A'}, "
                    f"última ocorrência: {_format_presentation_time(last_bad) or 'N/A'}. "
                    f"Os dados desses intervalos não devem ser utilizados sem validação prévia."
                )
            elif unknown_days:
                diag_text = "Foram identificados valores não mapeados pelo Digital Set. Verifique a aba Valores_Desconhecidos."
            elif uncovered_days:
                diag_text = "Foram encontradas lacunas de dados no período. Verifique a aba Linha_do_Tempo para detalhamento."
            else:
                diag_text = "Análise realizada com 100% de cobertura e qualidade válida. Nenhum problema de coleta ou estado desconhecido foi detectado."

            rows.append(["Tag:", r.metadata.tag, "Descrição:", descriptor, "Digital Set:", digital_set] + [""] * (num_cols - 6))
            r_i = len(rows)
            cell_styles[(r_i, 1)] = header_lbl_style
            cell_styles[(r_i, 3)] = header_lbl_style
            cell_styles[(r_i, 5)] = header_lbl_style

            rows.append(["Início Janela:", p_start_local_str, "Fim Janela:", p_end_local_str, "Fuso Horário:", DEFAULT_PRESENTATION_TZ] + [""] * (num_cols - 6))
            r_i = len(rows)
            cell_styles[(r_i, 1)] = header_lbl_style
            cell_styles[(r_i, 3)] = header_lbl_style
            cell_styles[(r_i, 5)] = header_lbl_style

            rows.append(["Status Executivo:", status_exec, "", "", "", ""] + [""] * (num_cols - 6))
            r_i = len(rows)
            cell_styles[(r_i, 1)] = header_lbl_style
            st_bg = "D32F2F" if "BAD" in status_exec else ("ED6C02" if ("DESCONHECIDO" in status_exec or "PARCIAL" in status_exec) else "2E7D32")
            cell_styles[(r_i, 2)] = XlsxCellStyle(bold=True, bg_color=st_bg, font_color="FFFFFF", align="center", border=True)
            merges.append(f"B{r_i}:{get_column_letter(num_cols)}{r_i}")

            rows.append(["Diagnóstico Executivo:", diag_text] + [""] * (num_cols - 2))
            r_i = len(rows)
            cell_styles[(r_i, 1)] = header_lbl_style
            cell_styles[(r_i, 2)] = XlsxCellStyle(wrap_text=True, border=True, bg_color="F7FAFC")
            merges.append(f"B{r_i}:{get_column_letter(num_cols)}{r_i}")

        rows.append([""] * num_cols)

        # Section 2: Resumo Consolidado dos Problemas (RF-04)
        rows.append(["2. RESUMO CONSOLIDADO DOS PROBLEMAS DE QUALIDADE E COBERTURA"] + [""] * (num_cols - 1))
        r_sec2 = len(rows)
        merges.append(f"A{r_sec2}:{get_column_letter(num_cols)}{r_sec2}")
        cell_styles[(r_sec2, 1)] = section_style

        rows.append(["Tipo de Problema", "Dias Afetados", "Intervalos Contínuos", "Pontos PI Individuais", "Duração Total", "Primeira Ocorrência", "Última Ocorrência"] + [""] * (num_cols - 7))
        r_hdr2 = len(rows)
        for c_i in range(1, 8):
            cell_styles[(r_hdr2, c_i)] = tbl_hdr_style

        prob_rows_data = [
            ("BAD", len(bad_days), qs.bad_segment_count if qs else 0, qs.bad_events if qs else 0, format_duration(cov.bad_seconds if cov else 0), _format_presentation_time(qs.first_bad_timestamp if qs else None), _format_presentation_time(qs.last_bad_timestamp if qs else None)),
            ("UNKNOWN", len(unknown_days), qs.unknown_segment_count if qs else 0, len(da_first.unknown_value_statistics) if da_first else 0, format_duration(cov.unknown_seconds if cov else 0), _format_presentation_time(qs.longest_unknown_start if qs else None), _format_presentation_time(qs.longest_unknown_end if qs else None)),
            ("NULL", len(null_days), 0, 0, format_duration(cov.null_seconds if cov else 0), "—", "—"),
            ("UNCOVERED", len(uncovered_days), 0, 0, format_duration(cov.uncovered_seconds if cov else 0), "—", "—"),
        ]
        for p_type, d_af, int_cont, pts_ind, dur_t, f_occ, l_occ in prob_rows_data:
            rows.append([p_type, d_af, int_cont, pts_ind, dur_t, f_occ or "—", l_occ or "—"] + [""] * (num_cols - 7))
            r_p_i = len(rows)
            for c_i in range(1, 8):
                cell_styles[(r_p_i, c_i)] = tbl_val_style
                if p_type == "BAD":
                    cell_styles[(r_p_i, 1)] = XlsxCellStyle(bold=True, bg_color="D32F2F", font_color="FFFFFF", align="center", border=True)
                elif p_type == "UNKNOWN":
                    cell_styles[(r_p_i, 1)] = XlsxCellStyle(bold=True, bg_color="ED6C02", font_color="FFFFFF", align="center", border=True)
                elif p_type == "NULL":
                    cell_styles[(r_p_i, 1)] = XlsxCellStyle(bold=True, bg_color="616161", font_color="FFFFFF", align="center", border=True)
                elif p_type == "UNCOVERED":
                    cell_styles[(r_p_i, 1)] = XlsxCellStyle(bold=True, bg_color="7E57C2", font_color="FFFFFF", align="center", border=True)

        rows.append([""] * num_cols)

        # Section 3: Indicadores Operacionais Secundários
        rows.append(["3. INDICADORES OPERACIONAIS SECUNDÁRIOS"] + [""] * (num_cols - 1))
        r_sec3 = len(rows)
        merges.append(f"A{r_sec3}:{get_column_letter(num_cols)}{r_sec3}")
        cell_styles[(r_sec3, 1)] = section_style

        for r in digital_results:
            da = r.digital_analysis
            cov = da.coverage if da else None

            known_pct = f"{cov.known_pct:.1f}%" if cov else "0.0%"
            uncovered_pct = f"{cov.uncovered_pct:.1f}%" if cov else "0.0%"
            transitions = sum(t.count for t in da.transition_statistics) if da and da.transition_statistics else 0

            dominant = "N/A"
            if da and da.state_statistics:
                obs_states = [s for s in da.state_statistics if s.observed]
                if obs_states:
                    dom_st = max(obs_states, key=lambda x: x.duration_seconds)
                    dominant = f"{dom_st.state_name} ({dom_st.percentage_of_window:.1f}%)"

            warnings_cnt = len(da.diagnostic_warnings) if da else 0

            rows.append(["Cobertura Conhecida", "Período Uncovered", "Total Transições", "Estado Predominante", "Alertas/Warnings", "Duração Janela"] + [""] * (num_cols - 6))
            r_idx = len(rows)
            for c_idx in range(1, 7):
                cell_styles[(r_idx, c_idx)] = kpi_card_style

            rows.append([known_pct, uncovered_pct, str(transitions), dominant, str(warnings_cnt), format_duration(cov.window_seconds) if cov else "0s"] + [""] * (num_cols - 6))
            r_idx_val = len(rows)
            for c_idx in range(1, 7):
                cell_styles[(r_idx_val, c_idx)] = kpi_val_style

        rows.append([""] * num_cols)

        # Section 4: Linha do Tempo Diária (GRID HORIZONTAL DE FAIXA DUPLA)
        rows.append(["4. LINHA DO TEMPO DIÁRIA (GRID HORIZONTAL DE FAIXA DUPLA)"] + [""] * (num_cols - 1))
        r_sec4 = len(rows)
        merges.append(f"A{r_sec4}:{get_column_letter(num_cols)}{r_sec4}")
        cell_styles[(r_sec4, 1)] = section_style

        month_groups: list[tuple[str, int, int]] = []
        curr_m_label = ""
        curr_m_start_col = 2
        for col_idx, d_val in enumerate(days, 2):
            m_label = d_val.strftime("%B %Y").capitalize()
            if m_label != curr_m_label:
                if curr_m_label:
                    month_groups.append((curr_m_label, curr_m_start_col, col_idx - 1))
                curr_m_label = m_label
                curr_m_start_col = col_idx
        if curr_m_label:
            month_groups.append((curr_m_label, curr_m_start_col, len(days) + 1))

        row_m_labels = ["Mês / Ano"] + [""] * (num_cols - 1)
        rows.append(row_m_labels)
        r_m_hdr = len(rows)
        cell_styles[(r_m_hdr, 1)] = XlsxCellStyle(bold=True, bg_color="1A365D", font_color="FFFFFF", align="center", border=True)
        for m_label, c_start, c_end in month_groups:
            rows[r_m_hdr - 1][c_start - 1] = m_label
            cell_styles[(r_m_hdr, c_start)] = XlsxCellStyle(bold=True, bg_color="1A365D", font_color="FFFFFF", align="center", border=True)
            if c_end > c_start:
                merges.append(f"{get_column_letter(c_start)}{r_m_hdr}:{get_column_letter(c_end)}{r_m_hdr}")

        row_d_labels = ["Dia / Mês (dd/mm)"] + [d_val.strftime("%d/%m") for d_val in days] + [""] * (num_cols - len(days) - 1)
        rows.append(row_d_labels)
        r_d_hdr = len(rows)
        cell_styles[(r_d_hdr, 1)] = XlsxCellStyle(bold=True, bg_color="DEE2E6", align="center", border=True)
        for c_i in range(2, len(days) + 2):
            cell_styles[(r_d_hdr, c_i)] = XlsxCellStyle(bold=True, bg_color="DEE2E6", align="center", border=True)

        row_qu_labels = ["1. Qualidade / Problema (PI)"]
        for d_sum in daily_summaries:
            p_label = d_sum["primary"]
            if d_sum["primary"] == "BAD" and d_sum["is_partial"]:
                p_label = f"BAD ({format_duration(d_sum['bad_dur'])})"
            elif d_sum["primary"] == "OK" and d_sum["is_partial"]:
                p_label = f"OK ({format_duration(d_sum['known_dur'])})"
            elif d_sum["kind_count"] > 1:
                p_label = f"{d_sum['primary']} (+)"
            row_qu_labels.append(p_label)
        row_qu_labels += [""] * (num_cols - len(daily_summaries) - 1)
        rows.append(row_qu_labels)
        r_qu_row = len(rows)
        cell_styles[(r_qu_row, 1)] = XlsxCellStyle(bold=True, bg_color="EDF2F7", align="left", border=True)
        for c_i, d_sum in enumerate(daily_summaries, 2):
            prim = d_sum["primary"]
            bg_c = _QUALITY_COLORS.get(prim, "2E7D32")
            if prim == "OK":
                bg_c = _QUALITY_COLORS["GOOD"]
            cell_styles[(r_qu_row, c_i)] = XlsxCellStyle(bg_color=bg_c, font_color="FFFFFF", align="center", bold=True, border=True)

        row_op_labels = ["2. Estado Operacional (Digital Set)"]
        for d_sum in daily_summaries:
            row_op_labels.append(d_sum["pred_state"])
        row_op_labels += [""] * (num_cols - len(daily_summaries) - 1)
        rows.append(row_op_labels)
        r_op_row = len(rows)
        cell_styles[(r_op_row, 1)] = XlsxCellStyle(bold=True, bg_color="EDF2F7", align="left", border=True)
        for c_i, d_sum in enumerate(daily_summaries, 2):
            cell_styles[(r_op_row, c_i)] = XlsxCellStyle(bg_color=d_sum["pred_color"], font_color="FFFFFF", align="center", bold=True, border=True)

        rows.append([""] * num_cols)

        # Section 5: Legenda Interpretativa
        rows.append(["5. LEGENDA INTERPRETATIVA"] + [""] * (num_cols - 1))
        r_sec5 = len(rows)
        merges.append(f"A{r_sec5}:{get_column_letter(num_cols)}{r_sec5}")
        cell_styles[(r_sec5, 1)] = section_style

        rows.append(["Categoria", "Rótulo / Classificação", "Cor", "Descrição Semântica / Significado"] + [""] * (num_cols - 4))
        r_leg_hdr = len(rows)
        for c_i in range(1, 5):
            cell_styles[(r_leg_hdr, c_i)] = XlsxCellStyle(bold=True, bg_color="EDF2F7")

        legend_items = [
            ("Qualidade", "OK / GOOD", _QUALITY_COLORS["GOOD"], "Coleta válida e estado digital reconhecido."),
            ("Qualidade", "BAD", _QUALITY_COLORS["BAD"], "Leitura inválida ou falha técnica registrada no PIMS."),
            ("Qualidade", "UNKNOWN", _QUALITY_COLORS["UNKNOWN"], "Valor lido fora do Digital Set configurado."),
            ("Qualidade", "NULL", _QUALITY_COLORS["NULL"], "Ausência de valor (ponto nulo no banco)."),
            ("Qualidade", "UNCOVERED", _QUALITY_COLORS["UNCOVERED"], "Lacuna ou gap de dados no período."),
            ("Operação", "Estados do Digital Set", _STATE_COLORS[0], "Cor de referência do estado operacional predominante no dia."),
        ]
        for cat, label, color, desc in legend_items:
            rows.append([cat, label, "   ", desc] + [""] * (num_cols - 4))
            r_idx_item = len(rows)
            cell_styles[(r_idx_item, 3)] = XlsxCellStyle(bg_color=color)

        column_widths_map: dict[int, float] = {1: 55.0}
        for c_i in range(2, len(days) + 2):
            column_widths_map[c_i] = 14.0

        columns_list = [f"C{i}" for i in range(1, num_cols + 1)]

        return XlsxSheet(
            name="Visao_Geral",
            columns=columns_list,
            rows=rows,
            is_presentation=True,
            is_active=True,
            freeze_panes="B1",
            column_widths=column_widths_map,
            merges=merges,
            cell_styles=cell_styles,
        )


    def _resumo_digital(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "description", "point_type", "digital_set",
            "window_start", "window_end", "window_duration_seconds",
            "analysis_status", "analysis_status_description",
            "recorded_event_count",
            "state_count_configured", "state_count_observed",
            "transition_count",
            "known_pct", "bad_pct", "unknown_pct", "null_pct", "uncovered_pct",
            "known_duration_seconds", "bad_duration_seconds",
            "unknown_duration_seconds", "null_duration_seconds", "uncovered_duration_seconds",
            "distinct_unknown_value_count",
            "longest_bad_duration_seconds", "longest_bad_start", "longest_bad_end",
            "longest_unknown_duration_seconds", "longest_unknown_start", "longest_unknown_end",
            "first_recorded_timestamp", "last_recorded_timestamp",
            "seed_found", "seed_timestamp", "seed_raw_value", "seed_good",
            "seed_classification", "seed_age_seconds_at_window_start",
            "executive_diagnosis", "investigative_recommendation",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                rows.append([r.metadata.tag] + [None] * (len(columns) - 1))
                continue
            cov = da.coverage
            qs = da.quality_summary
            si = da.seed_info
            status_val = da.status.value
            status_desc = _STATUS_DESCRIPTIONS.get(status_val, "")
            diagnosis = _DIAGNOSIS_TEMPLATES.get(status_val, "")

            # Recommendation do primeiro warning
            recommendation = ""
            for dw in da.diagnostic_warnings:
                rec = _WARNING_RECOMMENDATIONS.get(dw.code, "")
                if rec:
                    recommendation = rec
                    break

            first_rec = None
            last_rec = None
            if da.classified_recorded_events:
                first_rec = da.classified_recorded_events[0].timestamp
                last_rec = da.classified_recorded_events[-1].timestamp

            rows.append([
                r.metadata.tag,
                r.metadata.descriptor or "",
                r.metadata.point_type,
                r.metadata.digital_set or "",
                multi.period_start,
                multi.period_end,
                round(cov.window_seconds, 4),
                status_val,
                status_desc,
                da.recorded_events_count,
                len(da.possible_states),
                sum(1 for s in da.state_statistics if s.observed),
                sum(t.count for t in da.transition_statistics),
                cov.known_pct,
                cov.bad_pct,
                cov.unknown_pct,
                cov.null_pct,
                cov.uncovered_pct,
                cov.known_seconds,
                cov.bad_seconds,
                cov.unknown_seconds,
                cov.null_seconds,
                cov.uncovered_seconds,
                len(da.unknown_value_statistics),
                qs.longest_bad_duration if qs else 0,
                qs.longest_bad_start.isoformat() if qs and qs.longest_bad_start else None,
                qs.longest_bad_end.isoformat() if qs and qs.longest_bad_end else None,
                qs.longest_unknown_duration if qs else 0,
                qs.longest_unknown_start.isoformat() if qs and qs.longest_unknown_start else None,
                qs.longest_unknown_end.isoformat() if qs and qs.longest_unknown_end else None,
                first_rec,
                last_rec,
                si.found if si else False,
                si.timestamp if si else None,
                si.raw_value if si else None,
                si.good if si else None,
                si.classification.value if si and si.classification else None,
                si.age_seconds_at_window_start if si else None,
                diagnosis,
                recommendation,
            ])
        return XlsxSheet(name="Resumo", columns=columns, rows=rows)

    def _qualidade_digital(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "point_type",
            "total_events", "good_events", "bad_events",
            "questionable_events", "substituted_events",
            "known_duration_seconds", "bad_duration_seconds",
            "unknown_duration_seconds", "null_duration_seconds", "uncovered_duration_seconds",
            "known_pct", "bad_pct", "unknown_pct", "null_pct", "uncovered_pct",
            "questionable_duration_seconds", "questionable_pct",
            "substituted_duration_seconds", "substituted_pct",
            "bad_segment_count", "longest_bad_duration_seconds",
            "longest_bad_start", "longest_bad_end",
            "first_bad_timestamp", "last_bad_timestamp",
            "unknown_segment_count", "longest_unknown_duration_seconds",
            "longest_unknown_start", "longest_unknown_end",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                rows.append([r.metadata.tag, r.metadata.point_type] + [None] * (len(columns) - 2))
                continue
            cov = da.coverage
            qs = da.quality_summary
            rows.append([
                r.metadata.tag,
                r.metadata.point_type,
                qs.total_events if qs else 0,
                qs.good_events if qs else 0,
                qs.bad_events if qs else 0,
                qs.questionable_events if qs else 0,
                qs.substituted_events if qs else 0,
                cov.known_seconds,
                cov.bad_seconds,
                cov.unknown_seconds,
                cov.null_seconds,
                cov.uncovered_seconds,
                cov.known_pct,
                cov.bad_pct,
                cov.unknown_pct,
                cov.null_pct,
                cov.uncovered_pct,
                cov.questionable_seconds,
                cov.questionable_pct,
                cov.substituted_seconds,
                cov.substituted_pct,
                qs.bad_segment_count if qs else 0,
                qs.longest_bad_duration if qs else 0,
                qs.longest_bad_start.isoformat() if qs and qs.longest_bad_start else None,
                qs.longest_bad_end.isoformat() if qs and qs.longest_bad_end else None,
                qs.first_bad_timestamp if qs else None,
                qs.last_bad_timestamp if qs else None,
                qs.unknown_segment_count if qs else 0,
                qs.longest_unknown_duration if qs else 0,
                qs.longest_unknown_start.isoformat() if qs and qs.longest_unknown_start else None,
                qs.longest_unknown_end.isoformat() if qs and qs.longest_unknown_end else None,
            ])
        return XlsxSheet(name="Qualidade", columns=columns, rows=rows)

    def _recorded_digital(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "timestamp", "raw_value",
            "resolved_code", "resolved_state",
            "classification", "good", "questionable", "substituted",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                continue
            for ev in da.classified_recorded_events:
                rows.append([
                    r.metadata.tag,
                    ev.timestamp,
                    ev.raw_value,
                    ev.resolved_code,
                    ev.resolved_state,
                    ev.classification.value,
                    ev.good,
                    ev.questionable,
                    ev.substituted,
                ])
        return XlsxSheet(name="Recorded", columns=columns, rows=rows)

    def _estados(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "state_code", "state_name", "observed",
            "entries_count", "exits_count", "segment_count",
            "duration_seconds", "duration_human", "percentage_of_window",
            "dwell_avg_seconds", "dwell_median_seconds",
            "dwell_min_seconds", "dwell_max_seconds",
            "first_seen", "last_seen",
            "longest_segment_start", "longest_segment_end",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                continue
            for ss in da.state_statistics:
                rows.append([
                    r.metadata.tag,
                    ss.state_code,
                    ss.state_name,
                    ss.observed,
                    ss.entries_count,
                    ss.exits_count,
                    ss.segment_count,
                    ss.duration_seconds,
                    format_duration(ss.duration_seconds) if ss.duration_seconds > 0 else "00s",
                    ss.percentage_of_window,
                    ss.dwell_avg_seconds,
                    ss.dwell_median_seconds,
                    ss.dwell_min_seconds,
                    ss.dwell_max_seconds,
                    ss.first_seen.isoformat() if ss.first_seen else None,
                    ss.last_seen.isoformat() if ss.last_seen else None,
                    ss.longest_segment_start.isoformat() if ss.longest_segment_start else None,
                    ss.longest_segment_end.isoformat() if ss.longest_segment_end else None,
                ])
        return XlsxSheet(name="Estados", columns=columns, rows=rows)

    def _transicoes(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "from_kind", "from_code", "from_name",
            "to_kind", "to_code", "to_name",
            "transition_count", "percentage_of_transitions",
            "first_transition", "last_transition",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                continue
            for ts in da.transition_statistics:
                rows.append([
                    r.metadata.tag,
                    ts.from_kind.value if ts.from_kind else None,
                    ts.from_code,
                    ts.from_name,
                    ts.to_kind.value if ts.to_kind else None,
                    ts.to_code,
                    ts.to_name,
                    ts.count,
                    ts.percentage_of_transitions,
                    ts.first_transition.isoformat() if ts.first_transition else None,
                    ts.last_transition.isoformat() if ts.last_transition else None,
                ])
        return XlsxSheet(name="Transicoes", columns=columns, rows=rows)

    def _linha_do_tempo(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "segment_start", "segment_end",
            "duration_seconds", "duration_human",
            "raw_value", "state_code", "state_name",
            "segment_kind", "good", "questionable", "substituted", "source",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                continue
            for seg in da.timeline_segments:
                rows.append([
                    r.metadata.tag,
                    seg.start.isoformat() if hasattr(seg.start, 'isoformat') else str(seg.start),
                    seg.end.isoformat() if hasattr(seg.end, 'isoformat') else str(seg.end),
                    seg.duration_seconds,
                    format_duration(seg.duration_seconds),
                    seg.raw_value,
                    seg.state_code,
                    seg.state_name,
                    seg.kind.value,
                    seg.good,
                    seg.questionable,
                    seg.substituted,
                    seg.source.value if seg.source else None,
                ])
        return XlsxSheet(name="Linha_do_Tempo", columns=columns, rows=rows)

    def _valores_desconhecidos(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "raw_value", "occurrences", "segment_count",
            "duration_seconds", "duration_human", "percentage_of_window",
            "first_seen", "last_seen", "sample_timestamp",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                continue
            for uv in da.unknown_value_statistics:
                rows.append([
                    r.metadata.tag,
                    uv.raw_value,
                    uv.occurrences,
                    uv.segment_count,
                    uv.duration_seconds,
                    format_duration(uv.duration_seconds),
                    uv.percentage_of_window,
                    uv.first_seen.isoformat() if uv.first_seen else None,
                    uv.last_seen.isoformat() if uv.last_seen else None,
                    uv.sample_timestamp,
                ])
        return XlsxSheet(name="Valores_Desconhecidos", columns=columns, rows=rows)

    def _resumo_diario(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "date",
            "known_pct", "bad_pct", "unknown_pct", "null_pct", "uncovered_pct",
            "transition_count",
            "dominant_state_code", "dominant_state_name", "dominant_state_pct",
            "distinct_states_observed", "distinct_unknown_values",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                continue
            for db in da.daily_summary:
                rows.append([
                    r.metadata.tag,
                    db.date,
                    db.known_pct,
                    db.bad_pct,
                    db.unknown_pct,
                    db.null_pct,
                    db.uncovered_pct,
                    db.transition_count,
                    db.dominant_state_code,
                    db.dominant_state_name,
                    db.dominant_state_pct,
                    db.distinct_states_observed,
                    db.distinct_unknown_values,
                ])
        return XlsxSheet(name="Resumo_Diario", columns=columns, rows=rows)

    def _digital_set(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = ["tag", "digital_set_name", "state_code", "state_name", "state_description"]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is None:
                continue
            for dse in da.digital_set_snapshot:
                rows.append([
                    r.metadata.tag,
                    r.metadata.digital_set or "",
                    dse.state_code,
                    dse.state_name,
                    dse.state_description,
                ])
        return XlsxSheet(name="Digital_Set", columns=columns, rows=rows)

    # ------------------------------------------------------------------
    # Mixed
    # ------------------------------------------------------------------

    def _project_mixed(self, multi: MultiTagAnalysisResult) -> list[XlsxSheet]:
        sheets: list[XlsxSheet] = []
        sheets.append(self._resumo_mixed(multi))
        sheets.append(self._qualidade_mixed(multi))
        sheets.append(self._recorded_mixed(multi))
        sheets.append(self._estatisticas(multi))
        sheets.append(self._interpolated_5m(multi))
        sheets.append(self._gaps(multi))
        sheets.append(self._spikes(multi))

        # Digital-specific sheets (only rows from digital tags)
        has_digital_states = any(
            r.digital_analysis is not None and r.digital_analysis.state_statistics
            for r in multi.results
        )
        has_digital_trans = any(
            r.digital_analysis is not None and r.digital_analysis.transition_statistics
            for r in multi.results
        )
        has_digital_timeline = any(
            r.digital_analysis is not None and r.digital_analysis.timeline_segments
            for r in multi.results
        )
        has_digital_unknown = any(
            r.digital_analysis is not None and r.digital_analysis.unknown_value_statistics
            for r in multi.results
        )
        has_digital_daily = any(
            r.digital_analysis is not None and r.digital_analysis.daily_summary
            for r in multi.results
        )
        has_digital_set = any(
            r.digital_analysis is not None and r.digital_analysis.digital_set_snapshot
            for r in multi.results
        )

        if has_digital_states:
            sheets.append(self._estados(multi))
        if has_digital_trans:
            sheets.append(self._transicoes(multi))
        if has_digital_timeline:
            sheets.append(self._linha_do_tempo(multi))
        if has_digital_unknown:
            sheets.append(self._valores_desconhecidos(multi))
        if has_digital_daily:
            sheets.append(self._resumo_diario(multi))
        if has_digital_set:
            sheets.append(self._digital_set(multi))

        sheets.append(self._erros_warnings(multi))
        sheets.append(self._metadados(multi))
        return sheets

    def _resumo_mixed(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        # Superset of resumo + resumo_digital columns
        columns = [
            "tag", "point_type", "descriptor", "eng_units",
            "period_start", "period_end", "quality_verdict",
            "digital_set", "window_duration_seconds",
            "analysis_status", "analysis_status_description",
            "recorded_event_count",
            "known_pct", "bad_pct", "unknown_pct", "null_pct", "uncovered_pct",
            "distinct_unknown_value_count",
            "executive_diagnosis", "investigative_recommendation",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is not None:
                cov = da.coverage
                diagnosis = _DIAGNOSIS_TEMPLATES.get(da.status.value, "")
                recommendation = ""
                for dw in da.diagnostic_warnings:
                    rec = _WARNING_RECOMMENDATIONS.get(dw.code, "")
                    if rec:
                        recommendation = rec
                        break
                rows.append([
                    r.metadata.tag,
                    r.metadata.point_type,
                    r.metadata.descriptor or "",
                    r.metadata.engineering_units or "",
                    multi.period_start,
                    multi.period_end,
                    "NÃO APLICÁVEL",
                    r.metadata.digital_set or "",
                    round(cov.window_seconds, 4),
                    da.status.value,
                    _STATUS_DESCRIPTIONS.get(da.status.value, ""),
                    da.recorded_events_count,
                    cov.known_pct,
                    cov.bad_pct,
                    cov.unknown_pct,
                    cov.null_pct,
                    cov.uncovered_pct,
                    len(da.unknown_value_statistics),
                    diagnosis,
                    recommendation,
                ])
            else:
                verdict = r.quality.verdict if r.quality else ""
                rows.append([
                    r.metadata.tag,
                    r.metadata.point_type,
                    r.metadata.descriptor or "",
                    r.metadata.engineering_units or "",
                    multi.period_start,
                    multi.period_end,
                    verdict,
                    None, None, None, None, None, None,
                    None, None, None, None, None, None, None, None,
                ])
        return XlsxSheet(name="Resumo", columns=columns, rows=rows)

    def _qualidade_mixed(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "point_type",
            "good_pct", "questionable_pct", "substituted_pct",
            "zero_pct", "zero_policy", "verdict",
            "total_events", "good_events", "bad_events",
            "known_duration_seconds", "bad_duration_seconds",
            "unknown_duration_seconds", "null_duration_seconds", "uncovered_duration_seconds",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is not None:
                cov = da.coverage
                qs = da.quality_summary
                rows.append([
                    r.metadata.tag,
                    r.metadata.point_type,
                    None, None, None, None, r.zero_policy_applied, "NÃO APLICÁVEL",
                    qs.total_events if qs else 0,
                    qs.good_events if qs else 0,
                    qs.bad_events if qs else 0,
                    cov.known_seconds,
                    cov.bad_seconds,
                    cov.unknown_seconds,
                    cov.null_seconds,
                    cov.uncovered_seconds,
                ])
            elif r.quality is not None:
                rows.append([
                    r.metadata.tag,
                    r.metadata.point_type,
                    r.quality.good_pct,
                    r.quality.questionable_pct,
                    r.quality.substituted_pct,
                    r.quality.zero_pct,
                    r.zero_policy_applied,
                    r.quality.verdict,
                    None, None, None, None, None, None, None, None,
                ])
            else:
                rows.append([
                    r.metadata.tag, r.metadata.point_type,
                    None, None, None, None, r.zero_policy_applied, "",
                    None, None, None, None, None, None, None, None,
                ])
        return XlsxSheet(name="Qualidade", columns=columns, rows=rows)

    def _recorded_mixed(self, multi: MultiTagAnalysisResult) -> XlsxSheet:
        columns = [
            "tag", "point_type",
            "timestamp", "value", "good", "questionable", "substituted",
            "raw_value", "resolved_code", "resolved_state", "classification",
        ]
        rows = []
        for r in multi.results:
            da = r.digital_analysis
            if da is not None:
                for ev in da.classified_recorded_events:
                    rows.append([
                        r.metadata.tag, r.metadata.point_type,
                        ev.timestamp, None, ev.good, ev.questionable, ev.substituted,
                        ev.raw_value, ev.resolved_code, ev.resolved_state,
                        ev.classification.value,
                    ])
        return XlsxSheet(name="Recorded", columns=columns, rows=rows)
