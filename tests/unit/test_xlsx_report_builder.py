from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from domain.analysis.services.xlsx_projection import XlsxSheet
from mcp_server.services.delivery.exceptions import ArtifactLimitExceededError
from mcp_server.services.delivery.xlsx_report_builder import XlsxReportBuilder


@pytest.fixture
def builder(tmp_path: Path) -> XlsxReportBuilder:
    return XlsxReportBuilder(
        temp_dir=str(tmp_path),
        max_rows=100,
        max_bytes=1_000_000,
        max_columns=10,
        max_cell_bytes=1024,
    )


def _make_sheet(name: str = "TestSheet", rows: int = 5) -> XlsxSheet:
    return XlsxSheet(
        name=name,
        columns=["col_a", "col_b", "col_c"],
        rows=[[f"val_{i}_{j}" for j in range(3)] for i in range(rows)],
    )


class TestWorkbookOpens:
    def test_workbook_valid(self, builder: XlsxReportBuilder) -> None:
        path = builder.build_xlsx([_make_sheet()])
        assert path.exists()
        assert path.read_bytes().startswith(b"PK")
        wb = load_workbook(str(path))
        assert len(wb.sheetnames) == 1
        wb.close()
        path.unlink(missing_ok=True)

    def test_mkstemp_fd_is_closed_once(self, builder: XlsxReportBuilder, monkeypatch, tmp_path: Path) -> None:
        path = tmp_path / "fd-check.xlsx"
        path.touch()
        closed: list[int] = []

        monkeypatch.setattr(
            "mcp_server.services.delivery.xlsx_report_builder.tempfile.mkstemp",
            lambda **_: (42, str(path)),
        )
        monkeypatch.setattr(
            "mcp_server.services.delivery.xlsx_report_builder.os.close",
            closed.append,
        )

        result = builder.build_xlsx([_make_sheet()])

        assert result == path
        assert closed == [42]
        result.unlink(missing_ok=True)


class TestSheetsCorrect:
    def test_multiple_sheets(self, builder: XlsxReportBuilder) -> None:
        sheets = [_make_sheet("S1"), _make_sheet("S2"), _make_sheet("S3")]
        path = builder.build_xlsx(sheets)
        wb = load_workbook(str(path))
        assert wb.sheetnames == ["S1", "S2", "S3"]
        wb.close()
        path.unlink(missing_ok=True)


class TestFormulaProtection:
    def test_formula_prefix_protected(self, builder: XlsxReportBuilder) -> None:
        sheet = XlsxSheet(
            name="Formula",
            columns=["data"],
            rows=[["=SUM(A1:A10)"], ["+addition"], ["-subtraction"]],
        )
        path = builder.build_xlsx([sheet])
        wb = load_workbook(str(path))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "'=SUM(A1:A10)"
        assert ws.cell(row=3, column=1).value == "'+addition"
        assert ws.cell(row=4, column=1).value == "'-subtraction"
        wb.close()
        path.unlink(missing_ok=True)


class TestSheetNameSanitized:
    def test_invalid_chars_removed(self, builder: XlsxReportBuilder) -> None:
        sheet = XlsxSheet(name="Bad:Name/*?", columns=["a"], rows=[["1"]])
        path = builder.build_xlsx([sheet])
        wb = load_workbook(str(path))
        assert "Bad_Name___" in wb.sheetnames[0]
        wb.close()
        path.unlink(missing_ok=True)

    def test_max_31_chars(self, builder: XlsxReportBuilder) -> None:
        long_name = "A" * 50
        sheet = XlsxSheet(name=long_name, columns=["a"], rows=[["1"]])
        path = builder.build_xlsx([sheet])
        wb = load_workbook(str(path))
        assert len(wb.sheetnames[0]) <= 31
        wb.close()
        path.unlink(missing_ok=True)


class TestDuplicateNames:
    def test_duplicates_resolved(self, builder: XlsxReportBuilder) -> None:
        sheets = [_make_sheet("Dup"), _make_sheet("Dup"), _make_sheet("Dup")]
        path = builder.build_xlsx(sheets)
        wb = load_workbook(str(path))
        assert len(wb.sheetnames) == 3
        assert len(set(wb.sheetnames)) == 3
        wb.close()
        path.unlink(missing_ok=True)


class TestLimits:
    def test_max_rows_exceeded(self, builder: XlsxReportBuilder) -> None:
        sheet = _make_sheet(rows=200)
        with pytest.raises(ArtifactLimitExceededError, match="row_count"):
            builder.build_xlsx([sheet])

    def test_max_bytes_exceeded(self, tmp_path: Path) -> None:
        builder = XlsxReportBuilder(
            temp_dir=str(tmp_path),
            max_rows=1_000_000,
            max_bytes=100,
            max_columns=50,
            max_cell_bytes=1024,
        )
        sheet = XlsxSheet(name="Big", columns=["a"], rows=[["x" * 200] for _ in range(10)])
        with pytest.raises(ArtifactLimitExceededError, match="file_bytes"):
            builder.build_xlsx([sheet])


class TestCleanup:
    def test_temp_file_removed_on_error(self, builder: XlsxReportBuilder) -> None:
        sheet = _make_sheet(rows=200)
        with pytest.raises(ArtifactLimitExceededError):
            builder.build_xlsx([sheet])
        remaining = list(builder._temp_dir.glob("*.xlsx"))
        assert len(remaining) == 0


class TestPresentationFeatures:
    def test_native_types_in_presentation_sheet(self, builder: XlsxReportBuilder) -> None:
        from datetime import datetime
        now = datetime(2026, 8, 1, 10, 0, 0)
        sheet = XlsxSheet(
            name="Visao_Geral_e_Timeline",
            columns=["tag", "count", "time"],
            rows=[["CPD_SECADOR", 42, now]],
            is_presentation=True,
        )
        path = builder.build_xlsx([sheet])
        wb = load_workbook(str(path))
        ws = wb["Visao_Geral_e_Timeline"]
        assert ws.cell(row=2, column=1).value == "CPD_SECADOR"
        assert ws.cell(row=2, column=2).value == 42
        assert isinstance(ws.cell(row=2, column=2).value, int)
        assert ws.cell(row=2, column=3).value == now
        wb.close()
        path.unlink(missing_ok=True)

    def test_structural_layout(self, builder: XlsxReportBuilder) -> None:
        sheet = XlsxSheet(
            name="Visao_Geral_e_Timeline",
            columns=["col1", "col2"],
            rows=[["val1", "val2"]],
            is_presentation=True,
            freeze_panes="A3",
            column_widths={1: 25.5, 2: 12.0},
            merges=["A1:B1"],
            is_active=True,
        )
        path = builder.build_xlsx([sheet])
        wb = load_workbook(str(path))
        ws = wb["Visao_Geral_e_Timeline"]
        assert ws.freeze_panes == "A3"
        assert ws.column_dimensions["A"].width == 25.5
        assert ws.column_dimensions["B"].width == 12.0
        assert "A1:B1" in [str(m) for m in ws.merged_cells.ranges]
        assert wb.active == ws
        wb.close()
        path.unlink(missing_ok=True)

    def test_pattern_fill_and_styles(self, builder: XlsxReportBuilder) -> None:
        from domain.analysis.services.xlsx_projection import XlsxCellStyle
        style = XlsxCellStyle(bg_color="00FF00", font_color="FFFFFF", bold=True, border=True)
        sheet = XlsxSheet(
            name="Visao_Geral_e_Timeline",
            columns=["col1"],
            rows=[["OK"]],
            is_presentation=True,
            cell_styles={(2, 1): style},
        )
        path = builder.build_xlsx([sheet])
        wb = load_workbook(str(path))
        ws = wb["Visao_Geral_e_Timeline"]
        cell = ws.cell(row=2, column=1)
        assert cell.fill.start_color.rgb in ("0000FF00", "00FF00")
        assert cell.font.bold is True
        wb.close()
        path.unlink(missing_ok=True)

