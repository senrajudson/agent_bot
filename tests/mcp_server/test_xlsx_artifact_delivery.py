from __future__ import annotations

from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

from openpyxl import load_workbook

from domain.analysis.services.xlsx_projection import XlsxSheet
from mcp_server.clients.google_drive_client import DriveUploadedFile, GoogleDriveClient
from mcp_server.services.delivery.contracts import ArtifactMetadata, RequestSummary
from mcp_server.services.delivery.drive_publisher import DefaultDrivePublisher
from mcp_server.services.delivery.manifest_builder import build_artifact_manifest
from mcp_server.services.delivery.xlsx_report_builder import XlsxReportBuilder


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_xlsx_artifact_pipeline_preserves_bytes_and_mime(tmp_path: Path) -> None:
    builder = XlsxReportBuilder(temp_dir=str(tmp_path))
    path = builder.build_xlsx(
        [XlsxSheet(name="Resumo", columns=["tag"], rows=[["LFI_TEST"]])]
    )

    try:
        file_bytes = path.read_bytes()
        assert file_bytes.startswith(b"PK")
        workbook = load_workbook(BytesIO(file_bytes))
        assert workbook.sheetnames == ["Resumo"]
        workbook.close()

        client = MagicMock(spec=GoogleDriveClient)
        client.upload_file.return_value = DriveUploadedFile(
            file_id="xlsx-test",
            name="report.xlsx",
            mime_type=XLSX_MIME,
            size=len(file_bytes),
            web_view_link="https://drive.google.com/file/xlsx-test/view",
            web_content_link=None,
            created_time="2026-08-12T00:00:00Z",
        )
        publisher = DefaultDrivePublisher(client)
        published = publisher.publish(
            file_bytes=file_bytes,
            filename="report.xlsx",
            mime_type=XLSX_MIME,
            app_properties={"source": "test"},
        )

        client.upload_file.assert_called_once_with(
            file_bytes=file_bytes,
            filename="report.xlsx",
            mime_type=XLSX_MIME,
            app_properties={"source": "test"},
        )
        assert published.mime_type == XLSX_MIME

        manifest = build_artifact_manifest(
            status="success",
            tool_name="generate_pi_tags_analysis_report",
            request_summary=RequestSummary(
                tool_name="generate_pi_tags_analysis_report",
                tags_requested=1,
                tags_processed=1,
            ),
            artifact_metadata=ArtifactMetadata(
                format="xlsx",
                filename=published.name,
                mime_type=published.mime_type,
                row_count=2,
                column_count=1,
                size_bytes=published.size_bytes,
                view_url=published.view_url,
            ),
        )
        assert manifest.to_dict()["artifact"]["mime_type"] == XLSX_MIME
        assert "text/csv" not in manifest.to_json()
    finally:
        path.unlink(missing_ok=True)
