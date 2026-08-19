from __future__ import annotations

import os
import re
import tempfile
from dataclasses import is_dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from mcp_server.services.delivery._csv_protection import apply_formula_protection
from mcp_server.services.delivery.exceptions import ArtifactLimitExceededError

_fill_cache: dict[str, PatternFill] = {}
_font_cache: dict[tuple[str | None, bool], Font] = {}
_align_cache: dict[tuple[str | None, bool], Alignment] = {}
_thin_border = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)


def _get_fill(color: str) -> PatternFill:
    if color not in _fill_cache:
        _fill_cache[color] = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return _fill_cache[color]


def _get_font(color: str | None, bold: bool) -> Font:
    key = (color, bold)
    if key not in _font_cache:
        kwargs: dict[str, Any] = {"bold": bold}
        if color:
            kwargs["color"] = color
        _font_cache[key] = Font(**kwargs)
    return _font_cache[key]


def _get_align(align: str | None, wrap_text: bool) -> Alignment:
    key = (align, wrap_text)
    if key not in _align_cache:
        kwargs: dict[str, Any] = {"vertical": "center"}
        if align:
            kwargs["horizontal"] = align
        if wrap_text:
            kwargs["wrap_text"] = True
        _align_cache[key] = Alignment(**kwargs)
    return _align_cache[key]


class XlsxReportBuilder:
    def __init__(
        self,
        *,
        temp_dir: str = "/tmp/agent_bot_mcp_artifacts",
        max_rows: int = 1_000_000,
        max_bytes: int = 104_857_600,
        max_columns: int = 250,
        max_cell_bytes: int = 32768,
    ) -> None:
        self._temp_dir = Path(temp_dir)
        self._max_rows = max_rows
        self._max_bytes = max_bytes
        self._max_columns = max_columns
        self._max_cell_bytes = max_cell_bytes

    def build_xlsx(
        self,
        sheets: list[Any],
    ) -> Path:
        self._validate_limits(sheets)
        sanitized = [self._sanitize_sheet(s) for s in sheets]
        deduped = self._dedup_sheet_names(sanitized)

        self._temp_dir.mkdir(parents=True, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(self._temp_dir))
        os.close(fd)
        path = Path(tmp_path)

        try:
            wb = Workbook()
            wb.remove(wb.active)

            for s in deduped:
                ws = wb.create_sheet(title=s.name[:31])
                self._write_sheet(ws, s)

            wb.save(str(path))

            size = path.stat().st_size
            if size > self._max_bytes:
                path.unlink(missing_ok=True)
                raise ArtifactLimitExceededError(
                    field="file_bytes",
                    limit=self._max_bytes,
                    actual=size,
                )

            return path

        except InvalidFileException as exc:
            path.unlink(missing_ok=True)
            raise ArtifactLimitExceededError(
                field="file_bytes",
                limit=self._max_bytes,
                actual=0,
            ) from exc
        except Exception:
            path.unlink(missing_ok=True)
            raise

    def _validate_limits(self, sheets: list[Any]) -> None:
        total_rows = sum(len(s.rows) for s in sheets)
        total_cols = max((len(s.columns) for s in sheets), default=0)
        if total_rows > self._max_rows:
            raise ArtifactLimitExceededError(
                field="row_count",
                limit=self._max_rows,
                actual=total_rows,
            )
        if total_cols > self._max_columns:
            raise ArtifactLimitExceededError(
                field="column_count",
                limit=self._max_columns,
                actual=total_cols,
            )

    def _write_sheet(self, ws: Any, sheet: Any) -> None:
        is_presentation = getattr(sheet, "is_presentation", False)
        cell_styles = getattr(sheet, "cell_styles", {})
        column_widths = getattr(sheet, "column_widths", {})
        merges = getattr(sheet, "merges", [])
        freeze_panes = getattr(sheet, "freeze_panes", None)
        is_active = getattr(sheet, "is_active", False)

        for col_idx, col_name in enumerate(sheet.columns, 1):
            val_header = str(col_name) if not is_presentation else col_name
            cell = ws.cell(
                row=1,
                column=col_idx,
                value=apply_formula_protection(str(val_header)) if isinstance(val_header, str) else val_header,
            )
            style = cell_styles.get((1, col_idx))
            if style:
                self._apply_style(cell, style)

        for row_idx, row in enumerate(sheet.rows, 2):
            for col_idx, value in enumerate(row, 1):
                if is_presentation:
                    if value is None:
                        val_to_write = None
                    elif isinstance(value, str):
                        cell_bytes = len(value.encode("utf-8"))
                        if cell_bytes > self._max_cell_bytes:
                            raise ArtifactLimitExceededError(
                                field="cell_bytes",
                                limit=self._max_cell_bytes,
                                actual=cell_bytes,
                            )
                        val_to_write = apply_formula_protection(value)
                    elif isinstance(value, (int, float, datetime, date)):
                        val_to_write = value
                    else:
                        cell_str = str(value)
                        cell_bytes = len(cell_str.encode("utf-8"))
                        if cell_bytes > self._max_cell_bytes:
                            raise ArtifactLimitExceededError(
                                field="cell_bytes",
                                limit=self._max_cell_bytes,
                                actual=cell_bytes,
                            )
                        val_to_write = apply_formula_protection(cell_str)
                else:
                    cell_str = str(value) if value is not None else ""
                    cell_bytes = len(cell_str.encode("utf-8"))
                    if cell_bytes > self._max_cell_bytes:
                        raise ArtifactLimitExceededError(
                            field="cell_bytes",
                            limit=self._max_cell_bytes,
                            actual=cell_bytes,
                        )
                    val_to_write = apply_formula_protection(cell_str)

                cell = ws.cell(row=row_idx, column=col_idx, value=val_to_write)
                style = cell_styles.get((row_idx, col_idx))
                if style:
                    self._apply_style(cell, style)

        if column_widths:
            for col_idx, width in column_widths.items():
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

        if merges:
            for merge_range in merges:
                ws.merge_cells(merge_range)

        if freeze_panes:
            ws.freeze_panes = freeze_panes

        if is_active:
            ws.parent.active = ws

    def _apply_style(self, cell: Any, style: Any) -> None:
        bg_color = getattr(style, "bg_color", None)
        font_color = getattr(style, "font_color", None)
        bold = getattr(style, "bold", False)
        align = getattr(style, "align", None)
        wrap_text = getattr(style, "wrap_text", False)
        border = getattr(style, "border", False)
        number_format = getattr(style, "number_format", None)

        if bg_color:
            cell.fill = _get_fill(bg_color)
        if font_color or bold:
            cell.font = _get_font(font_color, bold)
        if align or wrap_text:
            cell.alignment = _get_align(align, wrap_text)
        if border:
            cell.border = _thin_border
        if number_format:
            cell.number_format = number_format

    def _sanitize_sheet(self, sheet: Any) -> Any:
        safe_name = re.sub(r"[\\[\]:*?/]", "_", sheet.name)[:31]
        if is_dataclass(sheet):
            return replace(sheet, name=safe_name)
        return sheet.__class__(name=safe_name, columns=sheet.columns, rows=sheet.rows)

    def _dedup_sheet_names(self, sheets: list[Any]) -> list[Any]:
        seen: set[str] = set()
        result: list[Any] = []
        for s in sheets:
            name = s.name
            if name in seen:
                counter = 2
                while f"{name}_{counter}" in seen:
                    counter += 1
                name = f"{name}_{counter}"
                if is_dataclass(s):
                    s = replace(s, name=name)
                else:
                    s = s.__class__(name=name, columns=s.columns, rows=s.rows)
            seen.add(name)
            result.append(s)
        return result
